# app.py
# ------------------------------------------------------------
# Análisis MICMAC Interactivo (total = directo + indirecto)
# by Martín Pratto — versión revisada
# ------------------------------------------------------------

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import io
from openpyxl import load_workbook
from datetime import datetime

# Etiquetas más limpias de matplotlib
plt.rcParams.update({
    "axes.titlesize": 18, "axes.labelsize": 14,
    "xtick.labelsize": 10, "ytick.labelsize": 10
})

# ------------------------------------------------------------
# Configuración de página
# ------------------------------------------------------------
st.set_page_config(page_title="Análisis MICMAC Interactivo", layout="wide")

st.markdown("""
# Análisis MICMAC Interactivo  
by **Martín Pratto**
""")
st.markdown("""
Herramienta visual para el análisis estructural de variables usando el método MICMAC.

**Cómo usar**
1) Sube tu matriz MICMAC en Excel (variables como filas y columnas; primera columna = nombres de fila).  
2) Ajusta **α** (atenuación) y **K** (longitud de rutas).  
3) Explora **rankings**, **tablas Directo/Indirecto/Total**, **gráficos**, y descarga **Excel/PNG/Informe**.
""")

# ------------------------------------------------------------
# Utilidades MICMAC
# ------------------------------------------------------------
def ensure_square_from_df(df: pd.DataFrame) -> pd.DataFrame:
    """Ajusta un DataFrame a matriz cuadrada usando la intersección fila/columna, y fuerza numeric + NaN->0."""
    df = df.apply(pd.to_numeric, errors='coerce').fillna(0.0)
    common = df.index.intersection(df.columns)
    if len(common) < 3:
        raise ValueError("No encuentro suficiente intersección entre filas y columnas para formar una matriz cuadrada.")
    df = df.loc[common, common].copy()
    # Diagonal a 0
    np.fill_diagonal(df.values, 0.0)
    return df

def micmac_total(M: np.ndarray, alpha: float, K: int) -> np.ndarray:
    """Total = M + α·M² + α²·M³ + … + α^(K−1)·M^K (diagonal forzada a 0)."""
    M = M.astype(float)
    total = M.copy()
    Mk = M.copy()
    for k in range(2, K+1):
        Mk = Mk @ M
        total += (alpha**(k-1)) * Mk
    np.fill_diagonal(total, 0.0)
    return total

def first_stable_K(M: np.ndarray, alpha: float, K_values=range(2, 11)) -> int:
    """Devuelve el primer K para el cual el ranking por motricidad no cambia respecto del K anterior."""
    prev_order = None
    for K in K_values:
        mot = micmac_total(M, alpha, K).sum(axis=1)
        order = tuple(np.argsort(-mot))
        if prev_order is not None and order == prev_order:
            return K
        prev_order = order
    return max(K_values)

def try_import_adjusttext():
    try:
        from adjustText import adjust_text
        return adjust_text
    except Exception:
        return None

adjust_text = try_import_adjusttext()

# ------------------------------------------------------------
# Carga de Excel
# ------------------------------------------------------------
uploaded_file = st.file_uploader("Sube tu archivo Excel MICMAC:", type=["xlsx"])

if uploaded_file:
    # Permitir selección de hoja
    try:
        wb = load_workbook(uploaded_file, data_only=True)
        sheets = wb.sheetnames
        sheet = st.selectbox("Selecciona la hoja con la matriz:", options=sheets, index=0)
        # Leer con pandas ahora que sabemos la hoja
        uploaded_file.seek(0)  # reset puntero
        df_raw = pd.read_excel(uploaded_file, sheet_name=sheet, index_col=0)
    except Exception as e:
        st.error(f"No puedo leer el Excel: {e}")
        st.stop()

    # Limpiezas opcionales (comunes)
    if 'SUMA' in df_raw.columns:
        df_raw = df_raw.drop(columns=['SUMA'])

    # Matriz cuadrada robusta
    try:
        df = ensure_square_from_df(df_raw)
    except Exception as e:
        st.error(str(e))
        st.stop()

    nombres = df.index.tolist()
    M = df.values.astype(float)

    st.success(f"✅ Archivo cargado. Hoja: **{sheet}** — Variables detectadas: **{len(nombres)}**")

    # --------------------------------------------------------
    # Parámetros
    # --------------------------------------------------------
    with st.sidebar:
        st.header("Parámetros")
        alpha = st.slider("α (atenuación de rutas indirectas)", 0.1, 1.0, 0.5, step=0.05)
        autoK = st.checkbox("Elegir K automáticamente por estabilidad", value=True)
        if autoK:
            K_max = first_stable_K(M, alpha)
            st.info(f"K estable detectado: **{K_max}**")
        else:
            K_max = st.slider("K (longitud máxima de rutas)", 2, 10, 6)
        usar_mediana = st.checkbox("Cuadrantes con mediana (en lugar de media)", value=True)
        max_etiquetas = st.slider("Máx. etiquetas en gráficos (para legibilidad)", 10, 60, 30, step=5)

    # --------------------------------------------------------
    # Cálculos MICMAC
    # --------------------------------------------------------
    # Directas
    mot_dir = M.sum(axis=1)
    dep_dir = M.sum(axis=0)

    # Totales (directo + indirecto)
    M_tot = micmac_total(M, alpha, K_max)
    mot_tot = M_tot.sum(axis=1)
    dep_tot = M_tot.sum(axis=0)

    # Indirectas = total - directa
    mot_ind = mot_tot - mot_dir
    dep_ind = dep_tot - dep_dir

    # DataFrame completo (índice = nombres)
    df_all = pd.DataFrame({
        "Motricidad_directa":   mot_dir,
        "Motricidad_indirecta": mot_ind,
        "Motricidad_total":     mot_tot,
        "Dependencia_directa":  dep_dir,
        "Dependencia_indirecta":dep_ind,
        "Dependencia_total":    dep_tot
    }, index=nombres)

    # Ranking por Motricidad total
    order = np.argsort(-mot_tot)
    ranking_vars = [nombres[i] for i in order]
    df_rank = pd.DataFrame({
        "Posición": np.arange(1, len(nombres)+1),
        "Variable": ranking_vars,
        "Motricidad_total": mot_tot[order],
        "Motricidad_directa": mot_dir[order],
        "Motricidad_indirecta": mot_ind[order],
        "Dependencia_total": dep_tot[order],
        "Dependencia_directa": dep_dir[order],
        "Dependencia_indirecta": dep_ind[order],
    })

    st.header(f"Ranking de Motricidad Total (α={alpha}, K={K_max})")
    st.dataframe(df_rank, use_container_width=True)

    # --------------------------------------------------------
    # Gráfico 1: Mapa MICMAC (Total)
    # --------------------------------------------------------
    st.subheader("Mapa MICMAC (Total)")
    X = mot_tot   # eje X estándar: Motricidad
    Y = dep_tot   # eje Y estándar: Dependencia

    ref_x = np.median(X) if usar_mediana else np.mean(X)
    ref_y = np.median(Y) if usar_mediana else np.mean(Y)

    # Clasificación cuadrantes
    labels_cuadrante = []
    for xi, yi in zip(X, Y):
        if   xi >= ref_x and yi <  ref_y: labels_cuadrante.append("Determinantes")
        elif xi >= ref_x and yi >= ref_y: labels_cuadrante.append("Crítico/inestable")
        elif xi <  ref_x and yi >= ref_y: labels_cuadrante.append("Variables resultado")
        else:                              labels_cuadrante.append("Autónomas")

    color_map = {
        "Determinantes":      "#FF4444",
        "Crítico/inestable":  "#1166CC",
        "Variables resultado":"#66BBFF",
        "Autónomas":          "#FF9944",
    }
    colors = [color_map[c] for c in labels_cuadrante]

    fig1, ax1 = plt.subplots(figsize=(12, 9))
    sc = ax1.scatter(X, Y, c=colors, s=120, alpha=0.85, edgecolors='black', linewidth=1.0)
    ax1.axvline(ref_x, color='black', linestyle='--', linewidth=1.2, alpha=0.8)
    ax1.axhline(ref_y, color='black', linestyle='--', linewidth=1.2, alpha=0.8)

    # Etiquetado (limitado por max_etiquetas)
    # Tomamos top por motricidad_total y por dependencia_total para cubrir ambos ejes
    idx_top_mot = np.argsort(-X)[:max_etiquetas//2]
    idx_top_dep = np.argsort(-Y)[:max_etiquetas//2]
    idx_show = np.unique(np.concatenate([idx_top_mot, idx_top_dep]))
    texts = []
    for i in idx_show:
        texts.append(ax1.text(X[i], Y[i], f" {nombres[i]}", fontsize=9))
    if adjust_text and len(texts) > 0:
        adjust_text(texts, ax=ax1, arrowprops=dict(arrowstyle='-', color='gray', lw=0.6))

    ax1.set_xlabel("Motricidad (Total)")
    ax1.set_ylabel("Dependencia (Total)")
    ax1.set_title(f"MICMAC Total — α={alpha}, K={K_max} — {'Mediana' if usar_mediana else 'Media'} como cortes")

    # Leyenda
    handles = [
        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Determinantes"], markersize=10, label='Determinantes'),
        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Crítico/inestable"], markersize=10, label='Crítico/inestable'),
        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Variables resultado"], markersize=10, label='Variables resultado'),
        plt.Line2D([0],[0], marker='o', color='w', markerfacecolor=color_map["Autónomas"], markersize=10, label='Autónomas'),
    ]
    ax1.legend(handles=handles, loc='upper left', frameon=True)
    ax1.grid(True, alpha=0.25)
    st.pyplot(fig1)

    # Descarga PNG
    buf1 = io.BytesIO()
    fig1.savefig(buf1, format="png", dpi=300, bbox_inches='tight'); buf1.seek(0)
    st.download_button("📥 Descargar Mapa MICMAC (PNG)", buf1, "micmac_mapa_total.png", "image/png")

    # --------------------------------------------------------
    # Gráfico 2: Eje de Estrategia (con totales normalizados)
    # --------------------------------------------------------
    st.subheader("Gráfico del Eje de Estrategia (totales normalizados)")
    x_norm = X / (X.max() if X.max()!=0 else 1.0)
    y_norm = Y / (Y.max() if Y.max()!=0 else 1.0)
    dist = np.abs(y_norm - x_norm) / np.sqrt(2)
    strategic_scores = (x_norm + y_norm)/2 - dist  # "cercanía a la diagonal + magnitud"

    # Colores por score (percentiles)
    p80, p60, p40 = np.percentile(strategic_scores, [80, 60, 40])
    col_est = []
    for s in strategic_scores:
        if s > p80:   col_est.append("#CC0000")
        elif s > p60: col_est.append("#FF6600")
        elif s > p40: col_est.append("#3388BB")
        else:         col_est.append("#888888")

    sizes_est = 50 + 100*(strategic_scores - strategic_scores.min())/(
        (strategic_scores.max() - strategic_scores.min()) if strategic_scores.max()>strategic_scores.min() else 1.0
    )

    fig2, ax2 = plt.subplots(figsize=(11, 9))
    ax2.scatter(X, Y, c=col_est, s=sizes_est, alpha=0.85, edgecolors='black', linewidth=1.0)
    # Eje estratégico (diagonal de los máximos)
    ax2.plot([0, X.max()], [0, Y.max()], 'r--', lw=2, label='Eje de estrategia')

    # Etiquetas: top-N estratégicas
    idx_top_est = np.argsort(strategic_scores)[-min(max_etiquetas, len(nombres)):]
    texts2 = []
    for i in idx_top_est:
        texts2.append(ax2.text(X[i], Y[i], f" {nombres[i]}", fontsize=9))
    if adjust_text and len(texts2) > 0:
        adjust_text(texts2, ax=ax2, arrowprops=dict(arrowstyle='-', color='orange', lw=0.6))

    ax2.set_xlabel("Motricidad (Total)")
    ax2.set_ylabel("Dependencia (Total)")
    ax2.set_title("Eje de Estrategia — mayor valor cuando hay alta magnitud y cercanía a la diagonal")
    ax2.legend(loc='upper left')
    ax2.grid(True, alpha=0.25)
    st.pyplot(fig2)

    # Descarga PNG
    buf2 = io.BytesIO()
    fig2.savefig(buf2, format="png", dpi=300, bbox_inches='tight'); buf2.seek(0)
    st.download_button("📥 Descargar Eje de Estrategia (PNG)", buf2, "micmac_eje_estrategia_total.png", "image/png")

    # --------------------------------------------------------
    # Scatter Motricidad vs Ranking (Total)
    # --------------------------------------------------------
    st.subheader("Motricidad Total vs Ranking (scatter)")
    fig3, ax3 = plt.subplots(figsize=(12, 5))
    ax3.scatter(range(1, len(nombres)+1), mot_tot[order])
    for idx, var in enumerate(ranking_vars):
        ax3.text(idx+1, mot_tot[order][idx], var[:20], fontsize=8, ha='center', rotation=90, va='bottom')
    ax3.set_xlabel("Ranking")
    ax3.set_ylabel("Motricidad Total")
    ax3.set_title(f"Motricidad Total vs Ranking (α={alpha}, K={K_max})")
    ax3.grid(True, alpha=0.25)
    st.pyplot(fig3)

    buf3 = io.BytesIO()
    fig3.savefig(buf3, format="png", dpi=300, bbox_inches='tight'); buf3.seek(0)
    st.download_button("📥 Descargar Scatter Ranking (PNG)", buf3, "micmac_scatter_ranking.png", "image/png")

    # --------------------------------------------------------
    # Barras (Top-N por Motricidad total)
    # --------------------------------------------------------
    st.subheader("Barras — Top variables por Motricidad Total")
    n_top = st.slider("Número de variables a mostrar:", 5, min(len(nombres), 25), 15)
    
    df_top = df_rank.head(n_top)
    fig4, ax4 = plt.subplots(figsize=(14, 6))
    bars = ax4.bar(range(len(df_top)), df_top["Motricidad_total"], color='steelblue', alpha=0.8)
    ax4.set_xticks(range(len(df_top)))
    ax4.set_xticklabels(df_top["Variable"], rotation=45, ha='right')
    ax4.set_ylabel("Motricidad Total")
    ax4.set_title(f"Top {n_top} Variables por Motricidad Total")
    ax4.grid(True, alpha=0.25, axis='y')
    plt.tight_layout()
    st.pyplot(fig4)

    buf4 = io.BytesIO()
    fig4.savefig(buf4, format="png", dpi=300, bbox_inches='tight'); buf4.seek(0)
    st.download_button("📥 Descargar Gráfico de Barras (PNG)", buf4, "micmac_barras_top.png", "image/png")

    # --------------------------------------------------------
    # Heatmap (opcional)
    # --------------------------------------------------------
    st.subheader("Heatmap — Comparación Directa vs Total")
    df_heatmap = pd.DataFrame({
        "Motricidad_directa": mot_dir,
        "Motricidad_total": mot_tot,
        "Dependencia_directa": dep_dir,
        "Dependencia_total": dep_tot
    }, index=nombres)
    
    fig5, ax5 = plt.subplots(figsize=(12, 8))
    sns.heatmap(df_heatmap.T, annot=False, cmap='YlOrRd', ax=ax5, cbar_kws={'shrink': 0.8})
    ax5.set_title("Heatmap: Motricidad y Dependencia (Directa vs Total)")
    ax5.set_xlabel("Variables")
    ax5.set_ylabel("Métricas")
    plt.tight_layout()
    st.pyplot(fig5)

    buf5 = io.BytesIO()
    fig5.savefig(buf5, format="png", dpi=300, bbox_inches='tight'); buf5.seek(0)
    st.download_button("📥 Descargar Heatmap (PNG)", buf5, "micmac_heatmap.png", "image/png")

    # --------------------------------------------------------
    # Tabla de Variables Estratégicas
    # --------------------------------------------------------
    st.subheader("Variables Más Estratégicas")
    top_estrategicas_idx = np.argsort(strategic_scores)[-15:][::-1]
    df_estrategicas = pd.DataFrame({
        'Variable': [nombres[i] for i in top_estrategicas_idx],
        'Motricidad_Total': [mot_tot[i] for i in top_estrategicas_idx],
        'Dependencia_Total': [dep_tot[i] for i in top_estrategicas_idx],
        'Puntuación_Estratégica': [strategic_scores[i] for i in top_estrategicas_idx],
        'Clasificación': [labels_cuadrante[i] for i in top_estrategicas_idx]
    })
    st.dataframe(df_estrategicas, use_container_width=True)

    # --------------------------------------------------------
    # Descargas Excel
    # --------------------------------------------------------
    st.subheader("📊 Descargas Excel")
    
    # Excel completo
    output_excel = io.BytesIO()
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df_rank.to_excel(writer, sheet_name='Ranking', index=False)
        df_all.to_excel(writer, sheet_name='Datos_Completos', index=True)
        df_estrategicas.to_excel(writer, sheet_name='Variables_Estrategicas', index=False)
        
        # Matriz original
        df.to_excel(writer, sheet_name='Matriz_Original', index=True)
        
        # Matriz total
        df_matriz_total = pd.DataFrame(M_tot, index=nombres, columns=nombres)
        df_matriz_total.to_excel(writer, sheet_name='Matriz_Total', index=True)
    
    output_excel.seek(0)
    st.download_button(
        label="📥 Descargar Análisis Completo (Excel)",
        data=output_excel,
        file_name=f"micmac_analisis_completo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # --------------------------------------------------------
    # GENERADOR DE INFORME DE INTELIGENCIA
    # --------------------------------------------------------
    st.subheader("🎯 Generar Informe de Inteligencia")
    
    if st.button("📄 Generar Informe de Inteligencia", type="primary"):
        
        # Análisis automático
        top_5_motoras = ranking_vars[:5]
        top_3_estrategicas = [nombres[i] for i in np.argsort(strategic_scores)[-3:]][::-1]
        
        # Contar por cuadrante
        count_determinantes = labels_cuadrante.count('Determinantes')
        count_criticas = labels_cuadrante.count('Crítico/inestable')
        count_resultado = labels_cuadrante.count('Variables resultado')
        count_autonomas = labels_cuadrante.count('Autónomas')
        
        fecha_actual = datetime.now().strftime("%d de %B de %Y")
        
        informe_contenido = f"""# INFORME DE INTELIGENCIA ESTRATÉGICA
**Análisis Estructural MICMAC - Sistema Complejo**  
*Generado automáticamente • {fecha_actual}*

---

## RESUMEN EJECUTIVO

El análisis MICMAC realizado sobre **{len(nombres)} variables** del sistema revela patrones estructurales críticos para la toma de decisiones estratégicas. Con parámetros de configuración α={alpha} y K={K_max}, se identificaron **{count_criticas} variables críticas/inestables** y **{count_determinantes} variables determinantes** que requieren atención prioritaria.

**HALLAZGO PRINCIPAL:** Las variables **{top_3_estrategicas[0]}**, **{top_3_estrategicas[1]}** y **{top_3_estrategicas[2]}** emergen como los factores de mayor valor estratégico del sistema.

---

## ANÁLISIS DE VARIABLES MOTORAS

### Top 5 Variables con Mayor Influencia Sistémica:
1. **{top_5_motoras[0]}** - Motricidad: {mot_tot[order[0]]:.0f}
2. **{top_5_motoras[1]}** - Motricidad: {mot_tot[order[1]]:.0f}  
3. **{top_5_motoras[2]}** - Motricidad: {mot_tot[order[2]]:.0f}
4. **{top_5_motoras[3]}** - Motricidad: {mot_tot[order[3]]:.0f}
5. **{top_5_motoras[4]}** - Motricidad: {mot_tot[order[4]]:.0f}

---

## CLASIFICACIÓN SISTÉMICA

| Categoría | Cantidad | Porcentaje |
|-----------|----------|------------|
| **Variables Críticas/Inestables** | {count_criticas} | {count_criticas/len(nombres)*100:.1f}% |
| **Variables Determinantes** | {count_determinantes} | {count_determinantes/len(nombres)*100:.1f}% |
| **Variables Resultado** | {count_resultado} | {count_resultado/len(nombres)*100:.1f}% |
| **Variables Autónomas** | {count_autonomas} | {count_autonomas/len(nombres)*100:.1f}% |

---

## RECOMENDACIONES ESTRATÉGICAS

### PRIORIDAD ALTA
1. **Focalización en Variables Determinantes:** Concentrar recursos en las {count_determinantes} variables determinantes identificadas.
2. **Gestión de Variables Críticas:** Desarrollar planes de contingencia para las {count_criticas} variables crítico/inestables.

### PRIORIDAD MEDIA
3. **Monitoreo de Variables Resultado:** Establecer KPIs basados en las {count_resultado} variables resultado.
4. **Optimización del Eje Estratégico:** Priorizar inversión en las variables más estratégicas.

---

## INDICADORES CLAVE

- **Motricidad Concentrada:** {(mot_tot[order[0]]/mot_tot.sum()*100):.2f}% (Variable líder)
- **Ratio Variables Críticas:** {count_criticas/len(nombres):.3f}
- **Dependencia Media:** {dep_tot.mean():.2f}

---

*Informe generado automáticamente por Sistema MICMAC Interactivo v2.0*  
*© 2025 - Martín Pratto*
"""
        
        # Crear archivo de texto
        informe_bytes = informe_contenido.encode('utf-8')
        
        st.success("✅ Informe de Inteligencia generado exitosamente!")
        st.download_button(
            label="📄 Descargar Informe (TXT)",
            data=informe_bytes,
            file_name=f"informe_inteligencia_{datetime.now().strftime('%Y%m%d_%H%M')}.txt",
            mime="text/plain"
        )
        
        with st.expander("👁️ Vista Previa del Informe", expanded=True):
            st.markdown(informe_contenido)

else:
    st.info("Por favor suba una matriz Excel para comenzar.")

st.caption("Desarrollado para análisis académico. ©2025")
